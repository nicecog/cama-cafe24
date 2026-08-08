#!/usr/bin/env python3
"""미매핑 100클래스에 국가표준식품성분(FC10.4) 코드를 안전하게 채운다.

식약처 D/P 코드가 없는 과일·유제품·흰밥 등은 검수된 A코드로만 보완한다.
"""
from __future__ import annotations

import csv
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
MAPPED = REPO / "docs" / "food_mvp_100_classes.mapped.csv"
CAND_OUT = REPO / "docs" / "food_fc10_mapping_candidates.csv"

# class_key → (food_code, food_name) — 수동 검수된 안전한 매핑만
SAFE_FC10: dict[str, tuple[str, str]] = {
    "white_rice": ("A013000A039a", "멥쌀, 백미, 밥"),
    "naengmyeon": ("A0060000001b", "메밀 냉면, 말린것, 삶은것"),
    "spaghetti": ("A0240020001b", "스파게티면, 말린것, 삶은것"),
    "fried_egg": ("J0040000009m", "달걀, 부침(달걀프라이)"),
    "scrambled_egg": ("J0060000009n", "달걀, 스크램블드에그"),
    "toast": ("A0210122739f", "빵, 식빵, 쇼트닝 첨가, 구운것"),
    "dumplings": ("S0260010007a", "만두, 고기 만두, 냉동"),
    "apple": ("H0500000000a", "사과, 생것"),
    "banana": ("H0370000000a", "바나나, 생것"),
    "orange": ("H0100020000a", "귤, 온주밀감, 생것"),
    "grape": ("H0980000000a", "포도, 생것"),  # fallback if missing
    "pear": ("H0380000000a", "배, 생것"),
    "persimmon": ("H0010010000a", "감, 단감, 생것"),
    "yogurt": ("M0080010009a", "요구르트, 액상"),
    "milk": ("M0090000009a", "우유"),
    "soy_milk": ("D0170000009a", "두유, 대두"),
    "americano": ("O027001158Pa", "커피, 아메리카노(2샷)"),
    "juice": ("H0720000009a", "오렌지 주스"),
    "soda": ("P0090090009a", "탄산 음료, 콜라"),
    "ion_drink": ("P0080002379a", "이온 음료, 레몬향"),
    "tofu": ("D0150000009a", "두부"),
    "sandwich": ("S0370020009a", "샌드위치, 닭고기"),
    "pork_cutlet": ("I0180000007a", "포크커틀릿, 냉동"),
    "bulgogi": ("I0190000000a", "돼지불고기, 생것"),
    "raw_fish": ("K0270000000a", "넙치(광어), 생것"),
}


def verify_codes_exist() -> dict[str, tuple[str, str]]:
    """부록2에서 실제 존재하는 코드만 반환. grape 등 대체 탐색."""
    from openpyxl import load_workbook

    xlsx = Path(r"D:/영양/식품성분표(10개정판).xlsx")
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = next(n for n in wb.sheetnames if n.startswith("부록2"))
    by_code: dict[str, str] = {}
    grape_candidates: list[tuple[str, str]] = []
    for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
        if i == 0 or not row or row[1] is None:
            continue
        code = str(row[1]).strip()
        name = str(row[2] or "").strip()
        if code and name:
            by_code[code] = name
            if name.startswith("포도") and "생것" in name and "잼" not in name:
                grape_candidates.append((code, name))
    wb.close()

    verified: dict[str, tuple[str, str]] = {}
    for key, (code, expected_name) in SAFE_FC10.items():
        if code in by_code:
            verified[key] = (code, by_code[code])
            continue
        if key == "grape" and grape_candidates:
            # prefer shortest generic name
            grape_candidates.sort(key=lambda x: len(x[1]))
            verified[key] = grape_candidates[0]
            continue
        print(f"WARN missing code for {key}: {code}")
    return verified


def main() -> None:
    verified = verify_codes_exist()
    rows = list(csv.DictReader(MAPPED.open(encoding="utf-8")))
    fieldnames = list(rows[0].keys())
    cand_rows: list[dict[str, str]] = []
    filled = 0

    for row in rows:
        key = row["class_key"]
        if key not in verified:
            continue
        code, name = verified[key]
        cand_rows.append(
            {
                "class_key": key,
                "name_ko": row.get("name_ko") or "",
                "candidate_code": code,
                "candidate_name": name,
                "candidate_score": "manual",
            }
        )
        if (row.get("food_code") or "").strip():
            continue
        row["food_code"] = code
        note = (row.get("notes") or "").strip()
        tag = f"fc10:{name}"
        row["notes"] = f"{note}; {tag}".strip("; ") if note else tag
        filled += 1
        print(f"  map {key} -> {code} | {name}")

    with MAPPED.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    with CAND_OUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["class_key", "name_ko", "candidate_code", "candidate_name", "candidate_score"],
        )
        writer.writeheader()
        writer.writerows(cand_rows)

    mapped_count = sum(1 for r in rows if (r.get("food_code") or "").strip())
    print(f"auto-filled={filled}, total_with_food_code={mapped_count}/{len(rows)}")
    print(f"candidates -> {CAND_OUT}")


if __name__ == "__main__":
    main()
