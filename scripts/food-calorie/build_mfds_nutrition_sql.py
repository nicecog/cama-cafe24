#!/usr/bin/env python3
"""식약처 식품영양성분 DB 파일 -> deploy/sql/cafe24-nutrition-mfds-load.sql 생성.

식약처 공공데이터포털 "식품의약품안전처 식품영양성분DB" (CSV / XLSX) 를 입력으로 받아
cm_food_nutrition 적재 SQL 과, cm_food_class 매핑 후보 CSV 를 만든다.

사용 예:
    python scripts/food-calorie/build_mfds_nutrition_sql.py \
        --input "D:/data/식품영양성분DB_음식.csv" \
        --version MFDS-2026.1

    # 매핑 후보만 다시 뽑고 싶을 때
    python scripts/food-calorie/build_mfds_nutrition_sql.py --input ... --mapping-only

매핑(class_key <-> food_code)은 사람이 검토해야 하므로 자동 적용하지 않는다.
docs/food_mvp_100_classes.mapped.csv 의 food_code 컬럼을 채운 뒤
--apply-mapping 으로 UPDATE 문을 함께 생성한다.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
CLASS_CSV = REPO_ROOT / "docs" / "food_mvp_100_classes.mapped.csv"
OUT_SQL = REPO_ROOT / "deploy" / "sql" / "cafe24-nutrition-mfds-load.sql"
OUT_MAPPING = REPO_ROOT / "docs" / "food_mfds_mapping_candidates.csv"

# 식약처 파일은 배포 회차마다 컬럼명이 조금씩 달라진다. 패턴으로 찾는다.
COLUMN_PATTERNS: dict[str, list[str]] = {
    "food_code": [r"^식품\s*코드$", r"^식품코드$", r"^코드$", r"FOOD_CD"],
    "food_name": [r"^식품\s*명$", r"^식품명$", r"^식품이름$", r"FOOD_NM"],
    "kcal": [r"에너지", r"열량", r"칼로리", r"ENERC"],
    "carb_g": [r"탄수화물", r"CHOCDF"],
    "protein_g": [r"단백질", r"PROCNT"],
    "fat_g": [r"지방(?!산)", r"FATCE"],
    "sugar_g": [r"당류", r"SUGAR"],
    "sodium_mg": [r"나트륨", r"^NA$"],
}

REQUIRED = ("food_code", "food_name", "kcal")


def normalize(text: str) -> str:
    text = unicodedata.normalize("NFKC", str(text or ""))
    return re.sub(r"\s+", "", text).strip()


def read_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise SystemExit(
                "XLSX 입력에는 openpyxl 이 필요합니다: pip install openpyxl"
            ) from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = [[("" if c is None else str(c)) for c in row] for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        rows = []
        for encoding in ("utf-8-sig", "cp949", "utf-8"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    rows = [row for row in csv.reader(handle)]
                break
            except UnicodeDecodeError:
                continue
        if not rows:
            raise SystemExit(f"인코딩을 판별할 수 없습니다: {path}")

    # 헤더 앞에 안내 행이 붙어 있는 파일이 있어 식품코드/식품명이 보이는 행을 헤더로 삼는다
    for index, row in enumerate(rows[:20]):
        joined = "".join(normalize(cell) for cell in row)
        if "식품코드" in joined or ("식품명" in joined and "에너지" in joined):
            return [normalize(c) for c in row], rows[index + 1:]
    return [normalize(c) for c in rows[0]], rows[1:]


def resolve_columns(header: list[str]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    for field, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            for index, name in enumerate(header):
                if index in resolved.values():
                    continue
                if re.search(pattern, name, flags=re.IGNORECASE):
                    resolved[field] = index
                    break
            if field in resolved:
                break

    missing = [field for field in REQUIRED if field not in resolved]
    if missing:
        raise SystemExit(
            "필수 컬럼을 찾지 못했습니다: "
            + ", ".join(missing)
            + f"\n헤더: {header}"
        )
    return resolved


def parse_number(value: str) -> str | None:
    text = re.sub(r"[^\d.\-]", "", str(value or ""))
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return f"{float(text):.2f}"
    except ValueError:
        return None


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def build_records(header: list[str], rows: list[list[str]]) -> list[dict[str, str | None]]:
    columns = resolve_columns(header)
    seen: set[str] = set()
    records: list[dict[str, str | None]] = []

    for row in rows:
        def cell(field: str) -> str:
            index = columns.get(field)
            if index is None or index >= len(row):
                return ""
            return str(row[index] or "").strip()

        food_code = cell("food_code")
        food_name = cell("food_name")
        kcal = parse_number(cell("kcal"))
        if not food_code or not food_name or kcal is None:
            continue
        if food_code in seen:
            continue
        seen.add(food_code)

        records.append({
            "food_code": food_code[:32],
            "food_name": food_name[:256],
            "kcal": kcal,
            "carb_g": parse_number(cell("carb_g")),
            "protein_g": parse_number(cell("protein_g")),
            "fat_g": parse_number(cell("fat_g")),
            "sugar_g": parse_number(cell("sugar_g")),
            "sodium_mg": parse_number(cell("sodium_mg")),
        })
    return records


def write_sql(
    records: list[dict[str, str | None]],
    version: str,
    mapping_updates: list[str],
    out_sql: Path = OUT_SQL,
) -> None:
    out_sql.parent.mkdir(parents=True, exist_ok=True)
    with out_sql.open("w", encoding="utf-8", newline="\n") as out:
        out.write("-- 식약처 식품영양성분 DB 적재 (자동 생성)\n")
        out.write(f"-- 생성 스크립트: scripts/food-calorie/build_mfds_nutrition_sql.py\n")
        out.write(f"-- nutrition_version: {version}\n")
        out.write(f"-- 행 수: {len(records)}\n")
        out.write("-- 적용: psql -f deploy/sql/cafe24-nutrition-mfds-load.sql\n")
        out.write("-- 선행: cafe24-nutrition-meal-log.sql\n\n")

        chunk_size = 500
        for start in range(0, len(records), chunk_size):
            chunk = records[start:start + chunk_size]
            out.write(
                "INSERT INTO public.cm_food_nutrition\n"
                "    (food_code, food_name, nutrition_version, kcal, carb_g, protein_g, fat_g, sugar_g, sodium_mg)\n"
                "VALUES\n"
            )
            values = []
            for record in chunk:
                values.append(
                    "    ("
                    + ", ".join([
                        sql_literal(record["food_code"]),
                        sql_literal(record["food_name"]),
                        sql_literal(version),
                        str(record["kcal"]),
                        record["carb_g"] or "NULL",
                        record["protein_g"] or "NULL",
                        record["fat_g"] or "NULL",
                        record["sugar_g"] or "NULL",
                        record["sodium_mg"] or "NULL",
                    ])
                    + ")"
                )
            out.write(",\n".join(values))
            out.write(
                "\nON CONFLICT (food_code, nutrition_version) DO UPDATE SET\n"
                "    food_name  = EXCLUDED.food_name,\n"
                "    kcal       = EXCLUDED.kcal,\n"
                "    carb_g     = EXCLUDED.carb_g,\n"
                "    protein_g  = EXCLUDED.protein_g,\n"
                "    fat_g      = EXCLUDED.fat_g,\n"
                "    sugar_g    = EXCLUDED.sugar_g,\n"
                "    sodium_mg  = EXCLUDED.sodium_mg;\n\n"
            )

        if mapping_updates:
            out.write("-- cm_food_class 매핑 반영 (docs/food_mvp_100_classes.mapped.csv 의 food_code)\n")
            for statement in mapping_updates:
                out.write(statement + "\n")
            out.write("\n")
            out.write(
                "-- 매핑된 클래스의 폴백 영양값을 정본으로 동기화한다.\n"
                "-- 앱 미리보기(catalog)와 서버 계산값의 차이를 줄이는 목적이다.\n"
                "UPDATE public.cm_food_class c\n"
                "   SET fb_kcal      = n.kcal,\n"
                "       fb_carb_g    = n.carb_g,\n"
                "       fb_protein_g = n.protein_g,\n"
                "       fb_fat_g     = n.fat_g,\n"
                "       updated_at   = now()\n"
                "  FROM public.cm_food_nutrition n\n"
                " WHERE n.food_code = c.food_code\n"
                f"   AND n.nutrition_version = {sql_literal(version)}\n"
                "   AND n.is_enabled = true;\n"
            )

        out.write("\nSELECT count(*) AS loaded FROM public.cm_food_nutrition WHERE nutrition_version = ")
        out.write(sql_literal(version) + ";\n")


def load_classes() -> list[dict[str, str]]:
    if not CLASS_CSV.is_file():
        raise SystemExit(f"클래스 목록이 없습니다: {CLASS_CSV}")
    with CLASS_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def build_mapping_updates(classes: list[dict[str, str]]) -> list[str]:
    statements = []
    for row in classes:
        food_code = (row.get("food_code") or "").strip()
        class_key = (row.get("class_key") or "").strip()
        if not food_code or not class_key:
            continue
        statements.append(
            "UPDATE public.cm_food_class SET food_code = "
            + sql_literal(food_code)
            + ", updated_at = now() WHERE class_key = "
            + sql_literal(class_key)
            + ";"
        )
    return statements


def write_mapping_candidates(classes: list[dict[str, str]], records: list[dict[str, str | None]]) -> None:
    names = [(record["food_code"], record["food_name"], normalize(record["food_name"])) for record in records]

    with OUT_MAPPING.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "class_id", "class_key", "name_ko", "current_food_code",
            "candidate_1_code", "candidate_1_name", "candidate_1_score",
            "candidate_2_code", "candidate_2_name", "candidate_2_score",
            "candidate_3_code", "candidate_3_name", "candidate_3_score",
        ])

        for row in classes:
            name_ko = (row.get("name_ko") or "").strip()
            # '흰밥(공기밥)' 처럼 괄호 보조설명은 매칭에서 제외한다
            base = normalize(re.sub(r"\(.*?\)", "", name_ko))
            scored = []
            for code, original, normalized in names:
                if not base:
                    continue
                score = SequenceMatcher(None, base, normalized).ratio()
                if base in normalized:
                    score += 0.35
                scored.append((score, code, original))
            scored.sort(reverse=True)

            output = [
                row.get("id", ""),
                row.get("class_key", ""),
                name_ko,
                (row.get("food_code") or "").strip(),
            ]
            for score, code, original in scored[:3]:
                output.extend([code, original, f"{min(score, 1.0):.2f}"])
            output.extend([""] * (13 - len(output)))
            writer.writerow(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", required=True, help="식약처 영양성분 CSV 또는 XLSX 경로")
    parser.add_argument("--version", default="MFDS-2026.1", help="nutrition_version 값")
    parser.add_argument(
        "--output",
        default=str(OUT_SQL),
        help="생성할 SQL 경로 (기본: deploy/sql/cafe24-nutrition-mfds-load.sql)",
    )
    parser.add_argument("--apply-mapping", action="store_true",
                        help="mapped.csv 의 food_code 를 cm_food_class 에 반영하는 UPDATE 문을 포함")
    parser.add_argument("--mapping-only", action="store_true",
                        help="매핑 후보 CSV 만 생성하고 SQL 은 만들지 않음")
    parser.add_argument("--skip-mapping-candidates", action="store_true",
                        help="매핑 후보 CSV 생성을 건너뛴다 (대용량 적재용)")
    args = parser.parse_args()

    source = Path(args.input)
    if not source.is_file():
        raise SystemExit(f"입력 파일이 없습니다: {source}")

    header, rows = read_rows(source)
    records = build_records(header, rows)
    if not records:
        raise SystemExit("적재할 행이 없습니다. 컬럼 매핑을 확인하세요.")

    if not args.skip_mapping_candidates:
        classes = load_classes()
        write_mapping_candidates(classes, records)
        print(f"매핑 후보: {OUT_MAPPING} ({len(classes)} 클래스)")
    else:
        classes = []

    if args.mapping_only:
        return

    mapping_updates = build_mapping_updates(load_classes()) if args.apply_mapping else []
    out_sql = Path(args.output)
    write_sql(records, args.version, mapping_updates, out_sql)
    print(f"적재 SQL: {out_sql} ({len(records)} 행, version={args.version})")
    if args.apply_mapping:
        print(f"매핑 UPDATE: {len(mapping_updates)} 건 포함")
    else:
        print("매핑 UPDATE 미포함 — 검토 후 --apply-mapping 으로 다시 생성하세요.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        print(f"FAILED: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
