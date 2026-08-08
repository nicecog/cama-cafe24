#!/usr/bin/env python3
"""식품성분표(10개정판) → cm_food_nutrition 적재 SQL 생성.

국가표준식품성분 Database 10.4 + 부록2(식품코드) 를 조인해
기존 식약처 D/P 코드와 다른 A코드 체계 데이터를 추가한다.

사용:
  python scripts/food-calorie/build_fc10_nutrition_sql.py
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = Path(r"D:/영양/식품성분표(10개정판).xlsx")
OUT_SQL = REPO / "deploy" / "sql" / "cafe24-nutrition-mfds-fc10-load.sql"
VERSION = "FC10.4"
SHEET_DB = "국가표준식품성분 Database 10.4"
SHEET_CODES = "부록2)식품코드,국문명,영문명,학명 정보 "


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def parse_number(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip().replace(",", "")
    if not text or text in {"-", "—", "–", "tr", "Tr", "N/A", "na"}:
        return None
    # trace amounts sometimes marked as "tr" or "<0.1"
    if text.startswith("<"):
        text = text[1:]
    try:
        return f"{float(text):.2f}"
    except ValueError:
        return None


def load_code_map(path: Path) -> dict[str, tuple[str, str]]:
    """DB색인 → (식품코드, 식품명)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    # sheet name may have trailing space
    sheet_name = None
    for name in wb.sheetnames:
        if name.startswith("부록2"):
            sheet_name = name
            break
    if not sheet_name:
        raise SystemExit("부록2 시트를 찾지 못했습니다")
    ws = wb[sheet_name]
    mapping: dict[str, tuple[str, str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None or row[1] is None:
            continue
        index = str(row[0]).strip()
        code = str(row[1]).strip()
        name = str(row[2] or "").strip()
        if not index.isdigit() or not code or code == "-":
            continue
        mapping[index] = (code, name)
    wb.close()
    return mapping


def load_records(path: Path, codes: dict[str, tuple[str, str]]) -> list[dict[str, str | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_DB not in wb.sheetnames:
        raise SystemExit(f"{SHEET_DB} 시트가 없습니다: {wb.sheetnames}")
    ws = wb[SHEET_DB]

    records: list[dict[str, str | None]] = []
    seen: set[str] = set()
    # rows: 0=title, 1=header names, 2=units, then data
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        if not row or row[0] is None:
            continue
        index = str(row[0]).strip()
        if not index.isdigit():
            continue
        food_name_cell = str(row[3] or "").strip()
        kcal = parse_number(row[5])
        if kcal is None:
            continue

        code_info = codes.get(index)
        if code_info:
            food_code, code_name = code_info
            food_name = code_name or food_name_cell
        else:
            # 코드 부록에 없으면 합성 코드로라도 보관해 참조 가능하게 한다
            food_code = f"FC10-{index.zfill(5)}"
            food_name = food_name_cell
        if not food_name:
            continue
        if food_code in seen:
            continue
        seen.add(food_code)

        records.append(
            {
                "food_code": food_code[:32],
                "food_name": food_name[:256],
                "kcal": kcal,
                "carb_g": parse_number(row[10]),
                "protein_g": parse_number(row[7]),
                "fat_g": parse_number(row[8]),
                "sugar_g": parse_number(row[11]),
                "sodium_mg": parse_number(row[26]),
            }
        )
    wb.close()
    return records


def write_sql(records: list[dict[str, str | None]], out_sql: Path, version: str) -> None:
    out_sql.parent.mkdir(parents=True, exist_ok=True)
    with out_sql.open("w", encoding="utf-8", newline="\n") as out:
        out.write("-- 국가표준식품성분표 제10개정 (Database 10.4) 적재\n")
        out.write("-- 생성: scripts/food-calorie/build_fc10_nutrition_sql.py\n")
        out.write(f"-- nutrition_version: {version}\n")
        out.write(f"-- 행 수: {len(records)}\n\n")

        chunk_size = 500
        for start in range(0, len(records), chunk_size):
            chunk = records[start : start + chunk_size]
            out.write(
                "INSERT INTO public.cm_food_nutrition\n"
                "    (food_code, food_name, nutrition_version, kcal, carb_g, protein_g, fat_g, sugar_g, sodium_mg)\n"
                "VALUES\n"
            )
            values = []
            for record in chunk:
                values.append(
                    "    ("
                    + ", ".join(
                        [
                            sql_literal(record["food_code"]),
                            sql_literal(record["food_name"]),
                            sql_literal(version),
                            str(record["kcal"]),
                            record["carb_g"] or "NULL",
                            record["protein_g"] or "NULL",
                            record["fat_g"] or "NULL",
                            record["sugar_g"] or "NULL",
                            record["sodium_mg"] or "NULL",
                        ]
                    )
                    + ")"
                )
            out.write(",\n".join(values))
            out.write(
                "\nON CONFLICT (food_code, nutrition_version) DO UPDATE SET\n"
                "    food_name  = EXCLUDED.food_name,\n"
                "    kcal       = EXCLUDED.kcal,\n"
                "    carb_g     = EXCLUDED.carb_g,\n"
                "    protein_g  = EXCLUDED.protein_g,\n"
                "    fat_g      = EXCLUDED.fat_g,\n"
                "    sugar_g    = EXCLUDED.sugar_g,\n"
                "    sodium_mg  = EXCLUDED.sodium_mg;\n\n"
            )
        out.write(
            f"SELECT count(*) AS loaded FROM public.cm_food_nutrition "
            f"WHERE nutrition_version = {sql_literal(version)};\n"
        )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(OUT_SQL))
    parser.add_argument("--version", default=VERSION)
    args = parser.parse_args()

    source = Path(args.input)
    if not source.is_file():
        raise SystemExit(f"입력 파일이 없습니다: {source}")

    print("loading code map (부록2)...")
    codes = load_code_map(source)
    print(f"codes={len(codes)}")
    print("loading Database 10.4...")
    records = load_records(source, codes)
    print(f"records={len(records)}")
    if not records:
        raise SystemExit("적재할 행이 없습니다")

    out = Path(args.output)
    write_sql(records, out, args.version)
    print(f"wrote {out}")

    # sample for white rice related
    rice = [r for r in records if "백미" in (r["food_name"] or "") and "밥" in (r["food_name"] or "")]
    print(f"sample white-rice-like={len(rice)}")
    for r in rice[:5]:
        print(f"  {r['food_code']} | {r['food_name']} | {r['kcal']} kcal")


if __name__ == "__main__":
    main()
